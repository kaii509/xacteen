import time
import pytest
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException
from appium.webdriver.common.touch_action import TouchAction

# Constants
APP_PACKAGE = "mn.xacbank.teen"
EXCEL_PATH = "/Users/khus1en/Documents/Internship Xac/XacTeen/tests/credentials.xlsx"
icon = chr(0xF002)   # 

@pytest.fixture(scope="module")
def sheet():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    # Set headers
    ws.cell(row=1, column=1).value = "Username"
    ws.cell(row=1, column=2).value = "Password"
    ws.cell(row=1, column=3).value = "Bank"
    ws.cell(row=1, column=4).value = "Account"
    ws.cell(row=1, column=5).value = "PIN"
    ws.cell(row=1, column=6).value = "Result"
    yield ws
    wb.save(EXCEL_PATH)
    wb.close()

def is_logged_in(driver):
    try:
        driver.find_element(
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Сүүлийн гүйлгээнүүд")'
        )
        return True
    except:
        return False

def logout(driver):
    wait = WebDriverWait(driver, 5)
    try:
        # open profile
        wait.until(EC.element_to_be_clickable(
            (AppiumBy.ANDROID_UIAUTOMATOR,
             'new UiSelector().className("android.widget.ImageView").instance(3)'))
        ).click()
        # click logout
        wait.until(EC.element_to_be_clickable(
            (AppiumBy.ANDROID_UIAUTOMATOR,
             'new UiSelector().description("Апп-аас гарах")'))
        ).click()
        time.sleep(1)
    except Exception:
        pass

def ensure_login_screen(driver):
    wait = WebDriverWait(driver, 1)

    try:
        # CASE 1: Logged in → navigate to profile and log out
        profile_icon = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(3)'
        )))
        print("[i] Already logged in. Logging out.")
        profile_icon.click()

        logout_btn = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("Апп-аас гарах")'
        )))
        logout_btn.click()
        time.sleep(2)

        # Wait for login button to appear again
        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("Нэвтрэх")'
        )))
        return

    except Exception:
        print("[i] Not already logged in. Checking screen...")

    # CASE 2: Press sign in button if present
    try:
        sign_in = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("Нэвтрэх")'
        )))
        sign_in.click()
        time.sleep(1)
    except Exception:
        print("[!] Sign-in button not found, maybe already in login screen.")

    # CASE 3: Handle saved-user screen (e.g., greeting)
    try:
        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Сайн уу")'
        )))
        print("[i] Found saved user screen. Navigating to default login.")

        other_user_btn = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Өөр хэрэглэгч?")'
        )))
        other_user_btn.click()
        time.sleep(1)
    except:
        print("[i] Default login screen ready.")
    
def enter_pin(driver, pin):
    wait = WebDriverWait(driver, 5)
    for digit in str(pin):
        try:
            wait.until(EC.presence_of_element_located((
                AppiumBy.ANDROID_UIAUTOMATOR,
                f'new UiSelector().text("{digit}")'
            ))).click()
            time.sleep(0.2)
        except Exception:
            return False
    return True

def transaction(driver, bank, account, amount, pin, accnf):
    wait = WebDriverWait(driver, 5)
    accnf = False
    
    try:
        # 1) go to transfer
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Мөнгө илгээх")'
        ))).click()

        # 2) select bank

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            f'new UiSelector().text("{bank}")'
        ))).click()

        # 3) tap account field
        account_num = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            f'new UiSelector().text("Данс, утасны дугаар")'
        )))
        account_num.clear()
        account_num.send_keys(account)

        # Search button 
        
        selector = f'new UiSelector().text("{icon}")'
        search_btn = wait.until(EC.element_to_be_clickable((AppiumBy.ANDROID_UIAUTOMATOR, selector)))
        search_btn.click()

        if  wait.until(EC.presence_of_element_located((AppiumBy.ANDROID_UIAUTOMATOR,f'new UiSelector().text("Хайлтын үр дүн олдсонгүй. Дансны нэр гараас оруулах боломжтой."))'))):
            accnf = True
            return accnf
            
        driver.terminate_app(APP_PACKAGE)
        time.sleep(1)
        driver.activate_app(APP_PACKAGE)
        time.sleep(1)
            
           
        


    except Exception:
        return False

def test_l(driver, sheet):
    wait = WebDriverWait(driver, 3)
    driver.activate_app(APP_PACKAGE)
    row = 2

    while True:
        ensure_login_screen(driver)

        user = sheet.cell(row=row, column=1).value
        pwd  = sheet.cell(row=row, column=2).value
        bank = sheet.cell(row=row, column=3).value
        acc  = sheet.cell(row=row, column=4).value
        pin  = sheet.cell(row=row, column=5).value
        accnf = sheet.cell(row=row, column=6).value
        amt = 100
        print(bank)

        # stop when empty
        if not user or not pwd:
            break

        # perform login
        try:
            username_field = wait.until(EC.presence_of_element_located((
                AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().className("android.widget.EditText").instance(0)'
            )))
            username_field.clear()
            username_field.send_keys(user)
        except Exception as e:
            sheet.cell(row=row, column = 6 ).value = "failed"
            continue

        # Enter password
        try:
            password_field = wait.until(EC.presence_of_element_located((
                AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().className("android.widget.EditText").instance(1)'
            )))
            password_field.clear()
            password_field.send_keys(pwd)
        except Exception as e:
            sheet.cell(row=row, column =6 ).value = "failed"
            continue

        try:
            login_btn = wait.until(EC.element_to_be_clickable((AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().description("Нэвтрэх")')))
            login_btn.click()
        except:
            print("Login button click failed.")

        time.sleep(2)

        if not is_logged_in(driver):
            sheet.cell(row=row, column=6).value = "login failed"
            row += 1
            driver.terminate_app(APP_PACKAGE)
            time.sleep(1)
            driver.activate_app(APP_PACKAGE)
            time.sleep(1)
            continue

        # perform transaction
        if transaction(driver, bank, acc, amt, pin, accnf):
            sheet.cell(row=row, column=7).value = "transfer passed"
        else:
            sheet.cell(row=row, column=7).value = "transfer failed"

        # prepare for next iteration
        logout(driver)
        driver.terminate_app(APP_PACKAGE)
        time.sleep(1)
        driver.activate_app(APP_PACKAGE)
        time.sleep(1)
        row += 1
